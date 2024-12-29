import pymysql
import os
import glob
import pandas as pd
import openpyxl


def get_lot_ids_from_xlsx(directory):
    lot_ids = []
    # Get all .xlsx files in the directory
    xlsx_files = glob.glob(os.path.join(directory, '*.xlsx'))

    for xlsx_file in xlsx_files:
        try:
            # Load the workbook and get the 'Sheet1'
            workbook = openpyxl.load_workbook(xlsx_file, data_only=True)
            if 'Sheet1' in workbook.sheetnames:
                sheet = workbook['Sheet1']
                # Extract values from the first column, excluding the header
                for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
                    if row[0] is not None:  # Ignore empty cells
                        lot_ids.append(row[0])
            # 关闭excel
            workbook.close()
        except Exception as e:
            print(f"Error processing file {xlsx_file}: {e}")
    return lot_ids

def mark_db(db_config, lots):
    connection = pymysql.connect(**db_config)
    try:
        with connection.cursor() as cursor:
            for lot_id in lots:
                # 更新表格信息
                update_sql = f"""
                    UPDATE db_fgs_request SET 
                    status = '1',
                    set_time = NOW()
                    WHERE lot_id = '{lot_id}'
                    """
                cursor.execute(update_sql)
            connection.commit()
            print("Data update successfully.")
    except Exception as e:
        print(f"Error updating data: {e}")
    finally:
        connection.close()

def main():
    db_config_local = {
        'host': 'localhost',
        'user': 'remoteuser',
        'password': 'password',
        'database': 'modulemte',
        'charset': 'utf8mb4'
    }
    db_config_apply = {
        'host': '172.27.154.57',
        'user': 'remoteuser',
        'password': 'password',
        'database': 'modulemte',
        'charset': 'utf8mb4'
    }
    directory = r'C:\Users\Tengjun Zhao\Desktop'  # Replace with the directory path containing the .xlsx files
    lots = get_lot_ids_from_xlsx(directory)
    mark_db(db_config_local, lots)
    # 删除directory 下所有excel
    xlsx_files = glob.glob(os.path.join(directory, '*.xlsx'))
    for xlsx_file in xlsx_files:
        os.remove(xlsx_file)
        print(f"File {xlsx_file} deleted.")

if __name__ == '__main__':
    main()