import pandas as pd
import openpyxl
import pymysql
from datetime import datetime, timedelta


def adjust_hitems_download(file_path:str):
    # 选择要操作的工作表（sheet）
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # 将 'Sheet1' 替换为您实际的工作表名称
    # 复制合并单元格的列表
    merged_cells = list(sheet.merged_cells)
    # 1. 取消所有合并单元格
    for cell_range in merged_cells:
        sheet.unmerge_cells(cell_range.coord)
    # 2. 遍历所有单元格，如果为空值则等于上一行数据
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            if cell.value is None:
                cell.value = cell.offset(row=-1).value
    # 3. 保存更改
    workbook.save(file_path)
    # 关闭 Excel 文件
    workbook.close()


# 确定prime Yiled et搜索起始使时间
def start_time_in_prime(table:str, field:str):
    db_config = {
        'host': 'localhost',
        'user': 'remoteuser',
        'password': 'password',
        'database': 'cmsalpha'
    }
    # 建立数据库连接
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()
    sql = f'SELECT MAX({field}) from {table};'
    cursor.execute(sql)
    res = cursor.fetchone()
    cursor.close()
    connection.close()
    date_obj = datetime.strptime(res[0], "%Y-%m-%d")
    new_date = date_obj + timedelta(days=1)
    new_date_string = new_date.strftime("%Y-%m-%d")
    return new_date_string

if __name__ == "__main__":
    # adjust_hitems_download(file_path=r'D:\Python\RPA Common\ztj.xlsx')
    print(start_time_in_prime('db_primeyieldet', 'date_val'))

