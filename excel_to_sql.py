import os
import glob
import pymysql
from openpyxl import load_workbook
import re


def main():
    # 配置数据库连接
    db_config = {
        'host': '172.27.154.57',
        'user': 'remoteuser',
        'password': 'password',
        'database': 'cmsalpha'
    }

    # 建立数据库连接
    connection = pymysql.connect(**db_config)
    cursor = connection.cursor()

    # 指定文件夹路径
    folder_path = r'D:\Python\RPA Common'

    # 获取路径下所有的 .xlsx 文件
    xlsx_files = glob.glob(os.path.join(folder_path, '*.xlsx'))

    for xlsx_file in xlsx_files:
        workbook = load_workbook(filename=xlsx_file, read_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[5] == 'Totals':
                continue  # 跳过"F"列为'Totals'的行


            # 提取equip_id中的部分
            # equip_id = re.split(r'\s*-\s*', row[5])[0]
            # 执行数据库插入操作
            insert_query = """
            INSERT INTO db_primeyieldat ( model, `date_val`, process, product, lot_id, equip_id, test_cnt, good_cnt, fail_cnt, retest_cnt,
            yield, fail_rate, retest_rate)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s)
            ON DUPLICATE KEY UPDATE 
            test_cnt = VALUES(test_cnt),
            good_cnt = VALUES(good_cnt),
            fail_cnt = VALUES(fail_cnt),
            retest_cnt = VALUES(retest_cnt),
            yield = VALUES(yield),
            fail_rate = VALUES(fail_rate),
            retest_rate = VALUES(retest_rate)
            """
            insert_data = (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10],
            row[11], row[12])
            print (insert_data)
            cursor.execute(insert_query, insert_data)
        connection.commit()
        workbook.close()
        # os.remove(xlsx_file)

    # 关闭数据库连接
    cursor.close()
    connection.close()

if __name__ == "__main__":
    main()  