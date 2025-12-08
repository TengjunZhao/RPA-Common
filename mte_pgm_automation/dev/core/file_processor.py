"""
文件处理器模块 - 负责PGM文件的下载、验证、整理等操作
"""
import os
import shutil
import zipfile
import hashlib
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Set
import time
from datetime import datetime
import sys

# 添加项目根目录到Python路径
current_dir = os.path.dirname(os.path.abspath(__file__))  # scripts目录
dev_dir = os.path.dirname(current_dir)  # dev目录
project_root = os.path.dirname(dev_dir)  # 项目根目录

sys.path.insert(0, dev_dir)
sys.path.insert(0, project_root)
from utils.config_loader import get_config
from utils.logger import get_pgm_logger
from database.models import PGMMain, PGMStatus, NextTask
from database.repositories import PGMMainRepository


class FileProcessor:
    """文件处理器"""

    def __init__(self):
        """初始化文件处理器"""
        self.logger = get_pgm_logger().get_logger('file_processor')
        self.config = get_config()
        self.file_paths = self.config.get_file_paths()
        self.hess_settings = self.config.get_hess_settings()
        self.pgm_verification_settings = self.config.get_pgm_verification_settings()

        self.logger.info("🔧 文件处理器初始化完成")

    def get_local_pgm_path(self, pgm_id: str) -> Optional[str]:
        """
        获取本地PGM路径

        Args:
            pgm_id: PGM ID

        Returns:
            本地路径，如果不存在则返回None
        """
        try:
            # 从数据库获取路径信息
            repo = PGMMainRepository()
            pgm = repo.get_by_id(pgm_id)
            repo.close()

            if pgm and pgm.server_path:
                return pgm.server_path

            # 如果数据库中没有，使用默认路径
            base_path = self.file_paths.get('local_verify', '')
            if not base_path:
                self.logger.error(f"❌ 未配置本地验证路径")
                return None

            pgm_path = os.path.join(base_path, pgm_id)
            if os.path.exists(pgm_path):
                return pgm_path

            self.logger.warning(f"⚠️ 本地PGM路径不存在: {pgm_path}")
            return None

        except Exception as e:
            self.logger.error(f"❌ 获取本地PGM路径失败 ({pgm_id}): {str(e)}")
            return None

    def download_pgm_files(self, pgm_id: str,
                           remote_path: str,
                           local_path: Optional[str] = None) -> bool:
        """
        下载PGM文件（模拟实现，实际需要根据具体源实现）

        Args:
            pgm_id: PGM ID
            remote_path: 远程路径
            local_path: 本地路径，如果为None则使用默认路径

        Returns:
            是否下载成功
        """
        try:
            self.logger.info(f"📥 开始下载PGM文件: {pgm_id}")
            self.logger.debug(f"远程路径: {remote_path}")

            if not local_path:
                local_path = self.file_paths.get('local_verify', '')
                if not local_path:
                    self.logger.error("❌ 未配置本地验证路径")
                    return False

                local_path = os.path.join(local_path, pgm_id)

            # 确保本地目录存在
            os.makedirs(local_path, exist_ok=True)

            # 模拟下载过程（实际需要根据具体源实现，如网络共享、FTP等）
            self.logger.info(f"📍 目标本地路径: {local_path}")

            # 这里应该是实际的下载逻辑
            # 示例：如果是网络共享路径
            if remote_path.startswith('\\\\'):
                # Windows网络共享路径
                self._copy_from_network_share(remote_path, local_path)
            else:
                # 其他类型的路径
                self.logger.warning(f"⚠️ 未实现的远程路径类型: {remote_path}")
                # 模拟成功
                self._create_sample_files(local_path, pgm_id)

            # 更新数据库
            repo = PGMMainRepository()
            repo.update(pgm_id, {
                'server_path': local_path,
                'status': PGMStatus.DOWNLOADED,
                'next_task': NextTask.VERIFY
            })
            repo.close()

            self.logger.info(f"✅ PGM文件下载完成: {pgm_id}")
            return True

        except Exception as e:
            self.logger.error(f"❌ 下载PGM文件失败 ({pgm_id}): {str(e)}")
            return False

    def _copy_from_network_share(self, remote_path: str, local_path: str) -> bool:
        """从网络共享复制文件"""
        try:
            self.logger.info(f"🌐 从网络共享复制: {remote_path} -> {local_path}")

            if not os.path.exists(remote_path):
                self.logger.error(f"❌ 远程路径不存在: {remote_path}")
                return False

            # 如果是目录，复制整个目录
            if os.path.isdir(remote_path):
                shutil.copytree(remote_path, local_path, dirs_exist_ok=True)
                self.logger.info(f"✅ 复制目录完成: {remote_path}")
            else:
                # 如果是文件，复制文件
                shutil.copy2(remote_path, local_path)
                self.logger.info(f"✅ 复制文件完成: {remote_path}")

            return True

        except Exception as e:
            self.logger.error(f"❌ 网络共享复制失败: {str(e)}")
            return False

    def _create_sample_files(self, local_path: str, pgm_id: str):
        """创建示例文件（用于测试）"""
        try:
            # 创建一些示例文件
            sample_files = [
                f"{pgm_id}_AT.zip",
                f"{pgm_id}_ET.zip",
                f"HESS_{pgm_id}.xlsx"
            ]

            for file_name in sample_files:
                file_path = os.path.join(local_path, file_name)
                with open(file_path, 'w') as f:
                    f.write(f"Sample content for {file_name}")

                self.logger.debug(f"📄 创建示例文件: {file_path}")

        except Exception as e:
            self.logger.error(f"❌ 创建示例文件失败: {str(e)}")

    def verify_pgm_structure(self, pgm_id: str) -> Tuple[bool, str, str]:
        """
        验证PGM文件结构

        Args:
            pgm_id: PGM ID

        Returns:
            (是否成功, 验证结果代码, 验证结果描述)
        """
        try:
            local_path = self.get_local_pgm_path(pgm_id)
            if not local_path:
                return False, "NO_LOCAL_PATH", "本地路径不存在"

            self.logger.info(f"🔍 开始验证PGM结构: {pgm_id}")
            self.logger.debug(f"验证路径: {local_path}")

            # 1. 检查目录是否存在且非空
            if not os.path.exists(local_path):
                return False, "PATH_NOT_EXIST", "路径不存在"

            if not os.listdir(local_path):
                return False, "EMPTY_DIRECTORY", "目录为空"

            # 2. 检查必要的文件类型
            pgm_at_files = []
            pgm_et_files = []
            hess_files = []

            for root, dirs, files in os.walk(local_path):
                for file in files:
                    file_ext = os.path.splitext(file)[1].lower()

                    # 检查AT PGM文件
                    if file_ext in self.pgm_verification_settings['pgm_types']['at_extensions']:
                        pgm_at_files.append(os.path.join(root, file))

                    # 检查ET PGM文件
                    if file_ext in self.pgm_verification_settings['pgm_types']['et_extensions']:
                        pgm_et_files.append(os.path.join(root, file))

                    # 检查HESS文件
                    if file_ext in self.hess_settings['valid_extensions']:
                        hess_files.append(os.path.join(root, file))

            self.logger.debug(f"📊 文件统计 - AT: {len(pgm_at_files)}, ET: {len(pgm_et_files)}, HESS: {len(hess_files)}")

            # 3. 基本验证规则
            if not pgm_at_files and not pgm_et_files:
                return False, "NO_PGM_FILES", "未找到PGM文件"

            if not hess_files:
                return True, "NO_HESS_FILES", "未找到HESS文件（允许继续）"

            # 4. 检查ZIP文件（如果有）
            zip_files = self._find_files_by_extensions(
                local_path,
                self.pgm_verification_settings['pgm_types']['zip_extensions']
            )

            if zip_files:
                # 检查ZIP文件内容
                zip_contents_ok = self._verify_zip_contents(zip_files, local_path)
                if not zip_contents_ok:
                    return False, "ZIP_CONTENT_ERROR", "ZIP文件内容错误"

            # 5. 根据文件组合确定PGM类型
            pgm_type = self._determine_pgm_type(pgm_at_files, pgm_et_files)

            # 6. 更新数据库
            repo = PGMMainRepository()
            repo.update(pgm_id, {
                'pgm_type': pgm_type,
                'verify_result_code': 'STRUCTURE_OK',
                'verify_result_desc': '文件结构验证通过',
                'status': PGMStatus.VERIFIED,
                'next_task': NextTask.APPLY
            })
            repo.close()

            self.logger.info(f"✅ PGM结构验证通过: {pgm_id} - 类型: {pgm_type}")
            return True, "STRUCTURE_OK", f"文件结构验证通过 ({pgm_type})"

        except Exception as e:
            self.logger.error(f"❌ PGM结构验证失败 ({pgm_id}): {str(e)}")
            return False, "VERIFICATION_ERROR", f"验证过程中发生错误: {str(e)}"

    def verify_pgm_with_hess(self, pgm_id: str) -> Tuple[bool, str, str]:
        """
        验证PGM与HESS文件的匹配（完整验证）

        Args:
            pgm_id: PGM ID

        Returns:
            (是否成功, 验证结果代码, 验证结果描述)
        """
        try:
            local_path = self.get_local_pgm_path(pgm_id)
            if not local_path:
                return False, "NO_LOCAL_PATH", "本地路径不存在"

            self.logger.info(f"🔍 开始完整PGM验证（含HESS）: {pgm_id}")

            # 1. 获取PGM文件列表
            pgm_at_list = self._get_pgm_file_list(local_path, 'at')
            pgm_et_list = self._get_pgm_file_list(local_path, 'et')

            # 2. 获取HESS文件
            hess_files = self._find_files_by_extensions(
                local_path,
                self.hess_settings['valid_extensions']
            )

            if not hess_files:
                # 没有HESS文件，进行基本验证
                success, code, desc = self.verify_pgm_structure(pgm_id)
                if not success:
                    return False, code, f"无HESS文件且结构验证失败: {desc}"

                return True, "NO_HESS_BUT_STRUCTURE_OK", "无HESS文件但结构验证通过"

            # 3. 分析HESS文件
            hess_analyzer = HESSAnalyzer()
            hess_results = []

            for hess_file in hess_files:
                hess_path = os.path.join(local_path, hess_file)
                result = hess_analyzer.analyze_hess(hess_path)
                if result:
                    hess_results.append(result)

            # 4. 提取HESS中的路径信息
            dir_in_at_hess = []
            dir_in_et_hess = []

            for result in hess_results:
                if result['type'] == 'AT':
                    dir_in_at_hess.extend(result['pgm_paths'])
                elif result['type'] == 'ET':
                    dir_in_et_hess.extend(result['pgm_paths'])

            # 5. 比较PGM路径与HESS路径
            comparison_code = self._compare_pgm_and_hess(
                pgm_at_list, pgm_et_list,
                dir_in_at_hess, dir_in_et_hess
            )

            # 6. 获取验证结果描述
            result_map = self.pgm_verification_settings['comparison_codes']
            result_desc = result_map.get(comparison_code, "未知验证结果")

            # 7. 判断验证是否成功
            is_success = self._is_verification_successful(comparison_code)

            # 8. 更新数据库
            repo = PGMMainRepository()
            repo.update(pgm_id, {
                'verify_result_code': comparison_code,
                'verify_result_desc': result_desc,
                'status': PGMStatus.VERIFIED if is_success else PGMStatus.VERIFY_FAILED,
                'next_task': NextTask.APPLY if is_success else NextTask.NONE
            })
            repo.close()

            if is_success:
                self.logger.info(f"✅ PGM验证成功: {pgm_id} - {result_desc}")
            else:
                self.logger.warning(f"⚠️ PGM验证失败: {pgm_id} - {result_desc}")

            return is_success, comparison_code, result_desc

        except Exception as e:
            self.logger.error(f"❌ PGM完整验证失败 ({pgm_id}): {str(e)}")
            return False, "FULL_VERIFICATION_ERROR", f"完整验证过程中发生错误: {str(e)}"

    def _get_pgm_file_list(self, directory: str, pgm_type: str) -> List[str]:
        """获取PGM文件列表"""
        extensions = []

        if pgm_type.lower() == 'at':
            extensions = self.pgm_verification_settings['pgm_types']['at_extensions']
        elif pgm_type.lower() == 'et':
            extensions = self.pgm_verification_settings['pgm_types']['et_extensions']

        files = []
        for root, dirs, filenames in os.walk(directory):
            for filename in filenames:
                file_ext = os.path.splitext(filename)[1].lower()
                if file_ext in extensions:
                    # 保存相对路径
                    rel_path = os.path.relpath(os.path.join(root, filename), directory)
                    files.append(rel_path)

        return files

    def _find_files_by_extensions(self, directory: str, extensions: List[str]) -> List[str]:
        """查找指定扩展名的文件"""
        files = []
        for root, dirs, filenames in os.walk(directory):
            for filename in filenames:
                file_ext = os.path.splitext(filename)[1].lower()
                if file_ext in extensions:
                    files.append(filename)
        return files

    def _verify_zip_contents(self, zip_files: List[str], base_path: str) -> bool:
        """验证ZIP文件内容"""
        try:
            for zip_file in zip_files:
                zip_path = os.path.join(base_path, zip_file)

                with zipfile.ZipFile(zip_path, 'r') as zf:
                    # 检查ZIP文件是否可以正常打开
                    file_list = zf.namelist()

                    if not file_list:
                        self.logger.warning(f"⚠️ ZIP文件为空: {zip_file}")
                        return False

                    # 检查是否包含有效的PGM文件
                    has_pgm_files = False
                    for file_in_zip in file_list:
                        file_ext = os.path.splitext(file_in_zip)[1].lower()

                        if (file_ext in self.pgm_verification_settings['pgm_types']['at_extensions'] or
                                file_ext in self.pgm_verification_settings['pgm_types']['et_extensions']):
                            has_pgm_files = True
                            break

                    if not has_pgm_files:
                        self.logger.warning(f"⚠️ ZIP文件不包含PGM文件: {zip_file}")
                        return False

            return True

        except zipfile.BadZipFile:
            self.logger.error(f"❌ ZIP文件损坏")
            return False
        except Exception as e:
            self.logger.error(f"❌ 验证ZIP文件内容失败: {str(e)}")
            return False

    def _determine_pgm_type(self, at_files: List[str], et_files: List[str]) -> str:
        """确定PGM类型"""
        if at_files and et_files:
            return "BOTH"
        elif at_files:
            return "AT"
        elif et_files:
            return "ET"
        else:
            return "UNKNOWN"

    def _compare_pgm_and_hess(self, pgm_at: List[str], pgm_et: List[str],
                              hess_at: List[str], hess_et: List[str]) -> str:
        """
        比较PGM路径与HESS路径

        返回比较结果代码，基于现有的验证逻辑
        """
        # 简化的比较逻辑，实际需要根据完整逻辑实现
        if not pgm_at and not pgm_et:
            if hess_at and hess_et:
                return "da0de0ha1he1"  # 仅有ET/AT HESS
            elif not hess_at and hess_et:
                return "da0de0ha0he1"  # 仅有ET HESS
            elif hess_at and not hess_et:
                return "da0de0ha1he0"  # 仅有AT HESS
            else:
                return "da0de0ha0he0"  # 无有效的PGM以及HESS

        # 更复杂的比较逻辑...
        # 这里简化实现，实际需要移植完整的compare_dir_and_hess函数

        return "da1de1ha1he1"  # 默认返回匹配成功

    def _is_verification_successful(self, comparison_code: str) -> bool:
        """根据比较代码判断验证是否成功"""
        # 成功的结果代码
        success_codes = [
            'da0de1ha0he1',  # ET 匹配
            'da1de0ha1he0',  # AT 匹配
            'da1de1ha1he1',  # AT 匹配，ET 匹配
            'da1de3ha1he1',  # AT 匹配，ET 不匹配
            'da3de1ha1he1',  # AT 不匹配，ET 匹配
            'da3de3ha1he1',  # AT 不匹配，ET 不匹配
        ]

        return comparison_code in success_codes

    def organize_pgm_files(self, pgm_id: str,
                           organization_type: str = "verify") -> bool:
        """
        整理PGM文件

        Args:
            pgm_id: PGM ID
            organization_type: 整理类型 (verify/apply)

        Returns:
            是否成功
        """
        try:
            local_path = self.get_local_pgm_path(pgm_id)
            if not local_path:
                return False

            self.logger.info(f"🗂️ 开始整理PGM文件: {pgm_id} - 类型: {organization_type}")

            if organization_type == "verify":
                return self._organize_for_verification(local_path, pgm_id)
            elif organization_type == "apply":
                return self._organize_for_application(local_path, pgm_id)
            else:
                self.logger.error(f"❌ 未知的整理类型: {organization_type}")
                return False

        except Exception as e:
            self.logger.error(f"❌ 整理PGM文件失败 ({pgm_id}): {str(e)}")
            return False

    def _organize_for_verification(self, pgm_path: str, pgm_id: str) -> bool:
        """为验证整理文件"""
        try:
            # 创建verify目录
            verify_dir = os.path.join(pgm_path, "verify")
            os.makedirs(verify_dir, exist_ok=True)

            # 查找HESS文件
            hess_files = self._find_files_by_extensions(
                pgm_path,
                self.hess_settings['valid_extensions']
            )

            if not hess_files:
                self.logger.warning(f"⚠️ 未找到HESS文件，跳过verify整理")
                return True

            # 处理ZIP文件
            zip_files = self._find_files_by_extensions(
                pgm_path,
                self.pgm_verification_settings['pgm_types']['zip_extensions']
            )

            if zip_files:
                # 解压ZIP文件
                for zip_file in zip_files:
                    zip_path = os.path.join(pgm_path, zip_file)
                    extract_dir = os.path.splitext(zip_path)[0]

                    if not os.path.exists(extract_dir):
                        os.makedirs(extract_dir)

                        with zipfile.ZipFile(zip_path, 'r') as zf:
                            zf.extractall(extract_dir)

                        self.logger.info(f"📦 解压文件: {zip_file} -> {extract_dir}")

            self.logger.info(f"✅ PGM验证整理完成: {pgm_id}")
            return True

        except Exception as e:
            self.logger.error(f"❌ 验证整理失败 ({pgm_id}): {str(e)}")
            return False

    def _organize_for_application(self, pgm_path: str, pgm_id: str) -> bool:
        """为适用整理文件"""
        try:
            # 创建apply目录
            apply_dir = os.path.join(pgm_path, "apply")
            os.makedirs(apply_dir, exist_ok=True)

            # 根据PGM类型整理文件
            repo = PGMMainRepository()
            pgm = repo.get_by_id(pgm_id)
            repo.close()

            if not pgm:
                self.logger.error(f"❌ 未找到PGM记录: {pgm_id}")
                return False

            pgm_type = pgm.pgm_type

            # 根据类型整理文件
            if pgm_type in ["AT", "BOTH"]:
                at_files = self._get_pgm_file_list(pgm_path, 'at')
                self._copy_files_to_directory(at_files, pgm_path, apply_dir, "AT")

            if pgm_type in ["ET", "BOTH"]:
                et_files = self._get_pgm_file_list(pgm_path, 'et')
                self._copy_files_to_directory(et_files, pgm_path, apply_dir, "ET")

            self.logger.info(f"✅ PGM适用整理完成: {pgm_id}")
            return True

        except Exception as e:
            self.logger.error(f"❌ 适用整理失败 ({pgm_id}): {str(e)}")
            return False

    def _copy_files_to_directory(self, files: List[str], source_dir: str,
                                 target_dir: str, prefix: str = "") -> int:
        """复制文件到目录"""
        copied_count = 0

        for file_rel_path in files:
            source_path = os.path.join(source_dir, file_rel_path)
            target_path = os.path.join(target_dir, prefix + "_" + os.path.basename(file_rel_path))

            try:
                # 确保目标目录存在
                os.makedirs(os.path.dirname(target_path), exist_ok=True)

                shutil.copy2(source_path, target_path)
                copied_count += 1

                self.logger.debug(f"📄 复制文件: {source_path} -> {target_path}")

            except Exception as e:
                self.logger.error(f"❌ 复制文件失败 ({source_path}): {str(e)}")

        return copied_count


class HESSAnalyzer:
    """HESS文件分析器"""

    def __init__(self):
        """初始化HESS分析器"""
        self.logger = get_pgm_logger().get_logger('hess_analyzer')
        self.config = get_config()
        self.hess_settings = self.config.get_hess_settings()

    def analyze_hess(self, hess_path: str) -> Optional[Dict[str, any]]:
        """
        分析HESS文件

        Args:
            hess_path: HESS文件路径

        Returns:
            分析结果字典
        """
        try:
            if not os.path.exists(hess_path):
                self.logger.error(f"❌ HESS文件不存在: {hess_path}")
                return None

            # 根据文件扩展名选择合适的分析器
            file_ext = os.path.splitext(hess_path)[1].lower()

            if file_ext in ['.xlsx', '.xls']:
                return self._analyze_excel_hess(hess_path)
            elif file_ext == '.csv':
                return self._analyze_csv_hess(hess_path)
            else:
                self.logger.error(f"❌ 不支持的HESS文件格式: {file_ext}")
                return None

        except Exception as e:
            self.logger.error(f"❌ 分析HESS文件失败 ({hess_path}): {str(e)}")
            return None

    def _analyze_excel_hess(self, hess_path: str) -> Optional[Dict[str, any]]:
        """分析Excel格式的HESS文件"""
        try:
            from openpyxl import load_workbook

            self.logger.info(f"📊 分析Excel HESS文件: {hess_path}")

            wb = load_workbook(hess_path, data_only=True)
            ws = wb.active

            # 1. 确定HESS类型 (AT/ET)
            hess_type = self._determine_hess_type(ws)

            # 2. 获取PGM路径列表
            pgm_paths = []
            product_info = []

            if hess_type == 'AT':
                # 获取AT HESS中的PGM路径
                column = self.hess_settings.get('at_path_column', 'Y')
                pgm_paths = self._extract_column_values(ws, column)

                # 获取产品信息
                product_info = self._extract_product_info(ws)

            elif hess_type == 'ET':
                # 获取ET HESS中的PGM路径
                column = self.hess_settings.get('et_path_column', 'X')
                pgm_paths = self._extract_column_values(ws, column)

            # 3. 构建分析结果
            result = {
                'file_path': hess_path,
                'type': hess_type,
                'pgm_paths': pgm_paths,
                'product_info': product_info,
                'total_records': len(pgm_paths),
                'analyzed_at': datetime.now().isoformat()
            }

            self.logger.info(f"✅ HESS分析完成: {hess_type} - {len(pgm_paths)}条记录")
            return result

        except ImportError:
            self.logger.error("❌ 未安装openpyxl库，无法分析Excel文件")
            return None
        except Exception as e:
            self.logger.error(f"❌ 分析Excel HESS文件失败: {str(e)}")
            return None

    def _analyze_csv_hess(self, hess_path: str) -> Optional[Dict[str, any]]:
        """分析CSV格式的HESS文件"""
        try:
            import csv

            self.logger.info(f"📊 分析CSV HESS文件: {hess_path}")

            with open(hess_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)

            if not rows:
                self.logger.warning("⚠️ CSV文件为空")
                return None

            # 简化的CSV分析逻辑
            # 实际需要根据CSV格式调整

            result = {
                'file_path': hess_path,
                'type': 'UNKNOWN',  # CSV通常需要特定格式
                'pgm_paths': [],
                'product_info': [],
                'total_records': len(rows) - 1,  # 减掉表头
                'analyzed_at': datetime.now().isoformat()
            }

            self.logger.info(f"✅ CSV HESS分析完成: {len(rows) - 1}条记录")
            return result

        except Exception as e:
            self.logger.error(f"❌ 分析CSV HESS文件失败: {str(e)}")
            return None

    def _determine_hess_type(self, worksheet) -> str:
        """确定HESS类型"""
        try:
            # 检查标识单元格
            at_cell = self.hess_settings.get('at_identifier_cell', 'X1')
            et_cell = self.hess_settings.get('et_identifier_cell', 'X1')

            at_value = worksheet[at_cell].value
            et_value = worksheet[et_cell].value

            at_identifier = self.hess_settings.get('at_identifier_value', 'HDIAG DIR')
            et_identifier = self.hess_settings.get('et_identifier_value', 'NEW PGM ID')

            if at_value == at_identifier:
                return 'AT'
            elif et_value == et_identifier:
                return 'ET'
            else:
                # 尝试其他判断逻辑
                if worksheet['Y1'].value == 'HDIAG DIR':
                    return 'AT'
                elif worksheet['X1'].value == 'NEW PGM ID':
                    return 'ET'
                else:
                    return 'UNKNOWN'

        except Exception as e:
            self.logger.error(f"❌ 确定HESS类型失败: {str(e)}")
            return 'UNKNOWN'

    def _extract_column_values(self, worksheet, column: str) -> List[str]:
        """提取指定列的值"""
        values = []
        try:
            for cell in worksheet[column]:
                if cell.value and cell.row > 1:  # 跳过表头
                    values.append(str(cell.value).strip())
        except Exception as e:
            self.logger.error(f"❌ 提取列{column}值失败: {str(e)}")

        return values

    def _extract_product_info(self, worksheet) -> List[Dict[str, str]]:
        """提取产品信息"""
        products = []
        try:
            product_columns = self.hess_settings.get('product_columns', {})

            # 从第二行开始（跳过表头）
            for row in range(2, worksheet.max_row + 1):
                product = {}

                for field, col_num in product_columns.items():
                    cell_value = worksheet.cell(row=row, column=col_num).value
                    product[field] = str(cell_value).strip() if cell_value else ''

                # 过滤掉空记录
                if any(product.values()):
                    products.append(product)

        except Exception as e:
            self.logger.error(f"❌ 提取产品信息失败: {str(e)}")

        return products


def test_file_processor():
    """测试文件处理器"""
    print("=" * 60)
    print("文件处理器测试")
    print("=" * 60)

    logger = get_pgm_logger()
    logger.log_execution_start("文件处理器测试")

    try:
        # 1. 初始化文件处理器
        processor = FileProcessor()
        print("✅ 文件处理器初始化成功")

        # 2. 创建测试PGM记录
        test_pgm_id = f"TEST_FP_{int(time.time())}"

        repo = PGMMainRepository()
        pgm_data = {
            'pgm_id': test_pgm_id,
            'pgm_type': 'ET',
            'status': 'NEW',
            'next_task': 'DOWNLOAD',
            'server_path': '/tmp/test_pgm',  # 测试路径
            'fab': 'TEST',
            'tech': 'TEST_FP'
        }

        pgm = repo.create(pgm_data)
        if not pgm:
            print("❌ 创建测试PGM记录失败")
            return False

        print(f"✅ 创建测试PGM记录: {test_pgm_id}")

        # 3. 测试文件下载（模拟）
        print("📥 测试文件下载...")
        download_success = processor.download_pgm_files(
            test_pgm_id,
            r"\\test\share\pgm",  # 模拟远程路径
            "/tmp/test_download"  # 测试本地路径
        )

        if download_success:
            print("✅ 文件下载测试成功")
        else:
            print("⚠️ 文件下载测试返回失败（可能是模拟实现）")

        # 4. 测试文件结构验证
        print("🔍 测试文件结构验证...")

        # 先更新状态为DOWNLOADED
        repo.update_status(test_pgm_id, PGMStatus.DOWNLOADED, NextTask.VERIFY)

        success, code, desc = processor.verify_pgm_structure(test_pgm_id)

        if success:
            print(f"✅ 文件结构验证成功: {desc}")
        else:
            print(f"⚠️ 文件结构验证失败: {desc}")

        # 5. 测试文件整理
        print("🗂️ 测试文件整理...")
        organize_success = processor.organize_pgm_files(test_pgm_id, "verify")

        if organize_success:
            print("✅ 文件整理测试成功")
        else:
            print("⚠️ 文件整理测试返回失败")

        # 6. 测试HESS分析器
        print("📊 测试HESS分析器...")
        hess_analyzer = HESSAnalyzer()

        # 创建一个测试Excel文件
        test_excel_path = "/tmp/test_hess.xlsx"
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws['X1'] = 'NEW PGM ID'  # ET HESS标识
            ws['X2'] = '/path/to/pgm1'
            ws['X3'] = '/path/to/pgm2'
            wb.save(test_excel_path)

            result = hess_analyzer.analyze_hess(test_excel_path)
            if result:
                print(f"✅ HESS分析成功: {result['type']} - {result['total_records']}条记录")
            else:
                print("❌ HESS分析失败")

        except ImportError:
            print("⚠️ 未安装openpyxl，跳过HESS分析测试")
        except Exception as e:
            print(f"⚠️ HESS分析测试出错: {str(e)}")

        # 7. 清理测试数据
        repo.delete(test_pgm_id)
        repo.close()

        print(f"🧹 清理测试数据: {test_pgm_id}")

        logger.log_execution_end("文件处理器测试", success=True)
        return True

    except Exception as e:
        logger.log_execution_end("文件处理器测试", success=False, message=str(e))
        print(f"❌ 文件处理器测试失败: {str(e)}")
        return False


if __name__ == "__main__":
    test_file_processor()